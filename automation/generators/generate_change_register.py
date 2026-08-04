"""
Project Orion
Enterprise Change Register Generator

Sprint 7.2
Version 1.0
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    )
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


class ChangeRegisterGenerator:
    """Generates the Project Orion Enterprise Change Register workbook."""

    def __init__(self):
        self.workbook = Workbook()

        # Project Orion Theme
        self.primary_fill = PatternFill("solid", fgColor="6B0F1A")
        self.secondary_fill = PatternFill("solid", fgColor="D9EAD3")

        self.title_font = Font(
            name="Calibri",
            size=22,
            bold=True,
            color="FFFFFF"
        )

        self.heading_font = Font(
            name="Calibri",
            size=14,
            bold=True
        )

        self.normal_font = Font(
            name="Calibri",
            size=11
        )

        self.header_font = Font(
            name="Calibri",
            size=11,
            bold=True,
            color="FFFFFF",
        )

        self.thin_border = Border(
            left=Side(style="thin", color="B7B7B7"),
            right=Side(style="thin", color="B7B7B7"),
            top=Side(style="thin", color="B7B7B7"),
            bottom=Side(style="thin", color="B7B7B7"),
        )

        self.wrap_center = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        self.center = Alignment(horizontal="center")

    def generate(self):
        """Generate the workbook."""

        self.create_home_sheet()
        self.create_change_register_sheet()
        self.create_change_history_sheet()
        self.create_cab_sheet()
        self.create_checklist_sheet()
        self.create_reference_sheet()
        self.create_revision_sheet()
        self.create_dashboard_sheet()

        output = Path(
            "docs/engineering/change-management/PO-CHANGE-REGISTER_v1.0.xlsx"
        )

        output.parent.mkdir(parents=True, exist_ok=True)

        self.workbook.save(output)

        print(f"Workbook created: {output}")

    def create_home_sheet(self):
        """Create the Home sheet."""

        sheet = self.workbook.active
        sheet.title = "00_Home"

        # ---------- Title ----------
        sheet.merge_cells("A1:F1")

        title = sheet["A1"]
        title.value = "PROJECT ORION"
        title.font = self.title_font
        title.fill = self.primary_fill
        title.alignment = self.center

        # ---------- Subtitle ----------
        sheet.merge_cells("A2:F2")

        subtitle = sheet["A2"]
        subtitle.value = "Enterprise Change Management Suite"
        subtitle.font = self.heading_font

        # ---------- Document Control ----------
        sheet["A4"] = "Document"
        sheet["B4"] = "PO-CHANGE-REGISTER_v1.0"

        sheet["A5"] = "Version"
        sheet["B5"] = "1.0"

        sheet["A6"] = "Status"
        sheet["B6"] = "Draft"

        sheet["A7"] = "Sprint"
        sheet["B7"] = "7.2"

        sheet["A8"] = "Owner"
        sheet["B8"] = "Project Orion"

        # ---------- Purpose ----------
        sheet["A10"] = "Purpose"

        sheet["A11"] = (
            "This workbook supports enterprise change management, "
            "validation, governance, and audit readiness."
        )

        # ---------- Workbook Navigation ----------
        sheet["A14"] = "Workbook Navigation"

        navigation = [
            "01_Change_Register",
            "02_Change_History",
            "03_CAB_Meetings",
            "04_Implementation_Checklist",
            "05_Reference_Data",
            "06_Revision_History",
            "07_Dashboard",
        ]

        row = 15

        for item in navigation:
            sheet.cell(row=row, column=1).value = item
            row += 1

        # ---------- Column Widths ----------
        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 35
        sheet.column_dimensions["C"].width = 20
        sheet.freeze_panes = "A4"

    def create_change_register_sheet(self):
        """Create the primary Change Register worksheet."""

        sheet = self.workbook.create_sheet("01_Change_Register")

        headers = [
            "Change ID",
            "Change Title",
            "Change Category",
            "Change Type",
            "Status",
            "Risk Level",
            "Priority",
            "Linked CI ID",
            "Asset ID",
            "Requested By",
            "Assigned Engineer",
            "CAB Approval",
            "Planned Start",
            "Planned End",
            "Actual Start",
            "Actual End",
            "Rollback Plan",
            "Validation Status",
            "Comments",
        ]

        # ---------- Title ----------
        sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=len(headers),
        )

        title = sheet.cell(row=1, column=1)
        title.value = "PROJECT ORION — ENTERPRISE CHANGE REGISTER"
        title.font = self.title_font
        title.fill = self.primary_fill
        title.alignment = self.center

        sheet.row_dimensions[1].height = 32

        # ---------- Subtitle ----------
        sheet.merge_cells(
            start_row=2,
            start_column=1,
            end_row=2,
            end_column=len(headers),
        )

        subtitle = sheet.cell(row=2, column=1)
        subtitle.value = (
            "Governed tracking of infrastructure, security, "
            "application, and operational changes"
        )
        subtitle.font = self.heading_font
        subtitle.alignment = self.center

        sheet.row_dimensions[2].height = 24

        # ---------- Headers ----------
        header_row = 3

        for column_number, header in enumerate(headers, start=1):
            cell = sheet.cell(
                row=header_row,
                column=column_number,
                value=header,
            )

            cell.font = self.header_font
            cell.fill = self.primary_fill
            cell.alignment = self.wrap_center
            cell.border = self.thin_border

        sheet.row_dimensions[header_row].height = 36

        # Add one blank table row.
        for column_number in range(1, len(headers) + 1):
            cell = sheet.cell(row=4, column=column_number)
            cell.border = self.thin_border

        # ---------- Excel Table ----------
        last_column = get_column_letter(len(headers))

        table = Table(
            displayName="ChangeRegisterTable",
            ref=f"A{header_row}:{last_column}4",
        )

        table_style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        table.tableStyleInfo = table_style
        sheet.add_table(table)

        # ---------- Date Formatting ----------
        date_columns = {
            "M",  # Planned Start
            "N",  # Planned End
            "O",  # Actual Start
            "P",  # Actual End
        }

        for column_letter in date_columns:
            sheet[f"{column_letter}4"].number_format = (
                "yyyy-mm-dd hh:mm"
            )

        # ---------- Column Widths ----------
        column_widths = {
            "A": 16,
            "B": 34,
            "C": 20,
            "D": 16,
            "E": 16,
            "F": 14,
            "G": 14,
            "H": 16,
            "I": 14,
            "J": 20,
            "K": 22,
            "L": 16,
            "M": 20,
            "N": 20,
            "O": 20,
            "P": 20,
            "Q": 16,
            "R": 18,
            "S": 42,
        }

        for column_letter, width in column_widths.items():
            sheet.column_dimensions[column_letter].width = width

        # ---------- Worksheet Controls ----------
        sheet.freeze_panes = "A4"
        sheet.sheet_view.showGridLines = False

    def create_change_history_sheet(self):
        self.workbook.create_sheet("02_Change_History")

    def create_cab_sheet(self):
        self.workbook.create_sheet("03_CAB_Meetings")

    def create_checklist_sheet(self):
        self.workbook.create_sheet("04_Implementation_Checklist")

    def create_reference_sheet(self):
        """Create the reference data sheet."""

        sheet = self.workbook.create_sheet("05_Reference_Data")

        reference_data = {
             "A": ("Change Type", [
                 "Standard",
                "Normal",
                "Emergency"
            ]),
            "C": ("Status", [
                "Planned",
                "Approved",
                "In Progress",
                "Completed",
                "Cancelled"
            ]),
            "E": ("Priority", [
                "Low",
                "Medium",
                "High",
                "Critical"
            ]),
            "G": ("Risk Level", [
                "Low",
                "Medium",
                "High",
                "Critical"
            ]),
            "I": ("CAB Approval", [
                "Yes",
                "No",
                "N/A"
            ]),
            "K": ("Validation Status", [
                "Pending",
                "Passed",
                "Warning",
                "Failed"
            ]),
        }

        for column, (title, values) in reference_data.items():

            header = sheet[f"{column}1"]
            header.value = title
            header.font = self.header_font
            header.fill = self.primary_fill
            header.alignment = self.center
            header.border = self.thin_border

            for row, value in enumerate(values, start=2):
                cell = sheet[f"{column}{row}"]
                cell.value = value
                cell.border = self.thin_border

            sheet.column_dimensions[column].width = 18

        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False

    def create_revision_sheet(self):
        self.workbook.create_sheet("06_Revision_History")

    def create_dashboard_sheet(self):
        self.workbook.create_sheet("07_Dashboard")
       
if __name__ == "__main__":
    ChangeRegisterGenerator().generate()